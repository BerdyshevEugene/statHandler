import os

from dotenv import load_dotenv
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side


load_dotenv()
EXCEL_PATH = os.getenv("EXCEL_PATH", "/default/path/to/report.xlsx")
if not EXCEL_PATH:
    raise ValueError(
        "EXCEL_PATH не задан! Укажите путь к Excel-файлу через переменную окружения или .env"
    )
EXCEL_PATH_LAST_YEAR = os.getenv("EXCEL_PATH_LAST_YEAR", "/default/path/to/report.xlsx")
if not EXCEL_PATH:
    raise ValueError(
        "EXCEL_PATH за прошлый год не задан! Укажите путь к Excel-файлу через переменную окружения или .env"
    )
RABBITMQ_URL = os.getenv("RABBITMQ_URL", "amqp://guest:guest@localhost/")
QUEUE_NAME = os.getenv("STATSCRAPER", "statScraper")


CITY_ORDER = [
    "Санкт-Петербург",
    "Москва",
    "Новосибирск",
    "Ростов-на-Дону",
    "Нижний Новгород",
]
CITY_MAPPING = {
    "spb": "Санкт-Петербург",
    "msk": "Москва",
    "nsk": "Новосибирск",
    "rnd": "Ростов-на-Дону",
    "nn": "Нижний Новгород",
}
MONTH_NAMES = {
    1: "Январь",
    2: "Февраль",
    3: "Март",
    4: "Апрель",
    5: "Май",
    6: "Июнь",
    7: "Июль",
    8: "Август",
    9: "Сентябрь",
    10: "Октябрь",
    11: "Ноябрь",
    12: "Декабрь",
}
WEEKDAYS = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
HEADERS = [
    "ВСЕГО:",
    "Потеряно:",
    "Переведено:",
    "Успешно завершено:",
    "Клиенты, не дождавшиеся ответа, ждали в среднем:",
    "В среднем клиенты ждут:",
    "В среднем разговор длится:",
    "% потерь",
]

HEADER_FILL = PatternFill(start_color="8DB4E2", end_color="8DB4E2", fill_type="solid")
HEADER_FONT = Font(bold=True)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
METRIC_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
METRIC_ALIGNMENT = Alignment(horizontal="left", vertical="center")
DATA_ALIGNMENT = Alignment(horizontal="center", vertical="center")

THICK_BORDER = Border(
    left=Side(style="medium"),
    right=Side(style="medium"),
    top=Side(style="medium"),
    bottom=Side(style="medium"),
)
