import openpyxl
import os

from datetime import datetime
from datetime import timedelta
from openpyxl.utils import get_column_letter
import math

from app.config import (
    EXCEL_PATH_LAST_YEAR,
    CITY_ORDER,
    CITY_MAPPING,
    MONTH_NAMES,
    WEEKDAYS,
    HEADERS,
    HEADER_FILL,
    HEADER_FONT,
    HEADER_ALIGNMENT,
    METRIC_FILL,
    METRIC_ALIGNMENT,
    DATA_ALIGNMENT,
    THICK_BORDER,
)
from app.send_email import send_email_with_attachment
from logger.logger import setup_logger
from openpyxl.styles import Font, PatternFill

logger = setup_logger(module_name=__name__)


class ExcelProcessor:
    def __init__(self, file_path):
        self.file_path = file_path
        try:
            self.wb = openpyxl.load_workbook(file_path)
        except FileNotFoundError:
            self.wb = openpyxl.Workbook()
            if "Sheet" in self.wb.sheetnames:
                del self.wb["Sheet"]
            self.wb.save(file_path)

        # устанавливаем ширину столбцов
        self.column_widths = {
            "A": 54.14,
            "B": 18.29,
            "C": 11.43,
            "D": 14.57,
            "E": 17.86,
            "F": 19.71,
        }

    def find_data_section(self, ws, target_date: datetime) -> int:
        date_str = target_date.strftime("%d.%m.%Y")
        weekday = WEEKDAYS[target_date.weekday()]
        search_value = f"{date_str} ({weekday})"

        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row, 1).value
            if cell_value == search_value:
                return row
        return None

    def get_city_column(self, ws, row: int, city: str) -> int:
        city = city.lower().strip()
        for col in range(2, ws.max_column + 1):
            cell_value = ws.cell(row, col).value
            if cell_value:
                cell_city = str(cell_value).lower().strip()
                if city in cell_city:
                    return col
        return None

    def create_new_data_section(self, ws, date: datetime) -> int:
        last_row = ws.max_row

        # добавляет ТОЛЬКО ОДНУ пустую строку перед новой секцией
        if last_row > 0:
            # проверяем, что последняя строка не пустая
            if any(ws.cell(last_row, col).value for col in range(1, ws.max_column + 1)):
                ws.append([])
                last_row += 1

        date_row = last_row + 1
        date_str = date.strftime("%d.%m.%Y")
        weekday = WEEKDAYS[date.weekday()]
        ws.cell(date_row, 1).value = f"{date_str} ({weekday})"
        ws.cell(date_row, 1).fill = HEADER_FILL
        ws.cell(date_row, 1).font = HEADER_FONT
        ws.cell(date_row, 1).alignment = HEADER_ALIGNMENT
        ws.cell(date_row, 1).border = THICK_BORDER

        for col, city in enumerate(CITY_ORDER, start=2):
            ws.cell(date_row, col).value = city
            ws.cell(date_row, col).fill = HEADER_FILL
            ws.cell(date_row, col).font = HEADER_FONT
            ws.cell(date_row, col).alignment = HEADER_ALIGNMENT
            ws.cell(date_row, col).border = THICK_BORDER

        for i, metric in enumerate(HEADERS, start=1):
            cell = ws.cell(date_row + i, 1)
            cell.value = metric
            cell.fill = METRIC_FILL
            cell.alignment = METRIC_ALIGNMENT
            cell.border = THICK_BORDER

        return date_row

    def create_new_sheet(
        self, sheet_name: str
    ) -> openpyxl.worksheet.worksheet.Worksheet:
        ws = self.wb.create_sheet(sheet_name)
        logger.info(f"new list created {sheet_name}")

        for col_letter, width in self.column_widths.items():
            ws.column_dimensions[col_letter].width = width
        return ws

    def format_data_cell(self, cell, value, index: int):
        cell.border = THICK_BORDER
        cell.alignment = DATA_ALIGNMENT

        if index in [4, 5, 6]:
            cell.value = f"{value} сек."
        elif index == 7:
            try:
                num = float(value)
                if num == int(num):
                    cell.value = f"{int(num)}%"
                else:
                    cell.value = f"{num:.1f}%".replace(".", ",")
            except (ValueError, TypeError):
                cell.value = value
        else:
            cell.value = value

    def get_previous_year_same_weekday(self, date):
        prev_year_date = date.replace(year=date.year - 1)
        while prev_year_date.weekday() != date.weekday():
            prev_year_date += timedelta(days=1)
        return prev_year_date

    def get_month_name(self, month):
        return MONTH_NAMES[month]

    def insert_previous_year_block(self, ws, data_row, date):
        logger.info(
            f"[prev] starting insertion of previous year block for date: {date}"
        )
        prev_date = self.get_previous_year_same_weekday(date)
        prev_file = EXCEL_PATH_LAST_YEAR
        logger.info(f"[prev] prev year file: {prev_file}, date: {prev_date}")
        if not os.path.exists(prev_file):
            logger.warning(f"[prev] previous year file not found: {prev_file}")
            return
        prev_wb = openpyxl.load_workbook(prev_file)
        prev_wb_values = openpyxl.load_workbook(prev_file, data_only=True)
        prev_sheet_name = (
            self.get_month_name(prev_date.month) + " " + str(prev_date.year)
        )
        logger.info(f"[prev] prev year sheet: {prev_sheet_name}")
        if prev_sheet_name not in prev_wb.sheetnames:
            logger.warning(
                f"[prev] sheet {prev_sheet_name} not found in previous year file"
            )
            return
        prev_ws = prev_wb[prev_sheet_name]
        prev_ws_values = prev_wb_values[prev_sheet_name]
        prev_data_row = self.find_data_section(prev_ws, prev_date)
        logger.info(f"[prev] prev year data row: {prev_data_row}")
        if not prev_data_row:
            logger.warning(
                f"[prev] section for date {prev_date} not found in previous year file"
            )
            return
        # Определяем конец секции: первая полностью пустая строка после prev_data_row
        start_row = prev_data_row
        end_row = start_row
        while True:
            if all(
                prev_ws.cell(row=end_row, column=col).value in (None, "")
                for col in range(1, 7)
            ):
                logger.info(f"[prev] prev year section end at row: {end_row}")
                break
            end_row += 1
        ws.append([])
        logger.info(f"[prev] copying rows: {start_row}-{end_row - 1}")
        for src_row in range(start_row, end_row):
            values = []
            for col in range(1, 7):
                cell = prev_ws.cell(row=src_row, column=col)
                cell_value = cell.value
                if cell.data_type == "f":
                    cell_value = prev_ws_values.cell(row=src_row, column=col).value
                values.append(cell_value)
            logger.info(f"[prev] inserting row: {values}")
            ws.append(values)
            for col in range(1, 7):
                cell = ws.cell(row=ws.max_row, column=col)
                if src_row == start_row:
                    cell.fill = HEADER_FILL
                    cell.font = HEADER_FONT
                    cell.alignment = HEADER_ALIGNMENT
                elif col == 1:
                    cell.fill = METRIC_FILL
                    cell.font = Font(bold=False)
                    cell.alignment = METRIC_ALIGNMENT
                else:
                    cell.fill = PatternFill(fill_type=None)
                    cell.font = Font(bold=False)
                    cell.alignment = DATA_ALIGNMENT
                cell.border = THICK_BORDER
                if (
                    col != 1
                    and "% потерь"
                    in str(ws.cell(row=ws.max_row, column=1).value).lower()
                ):
                    val = cell.value
                    try:
                        num = float(val) * 100
                        if num == int(num):
                            cell.value = f"{int(num)}%"
                        else:
                            cell.value = (
                                f"{num:.1f}".replace(".", ",").rstrip("0").rstrip(",")
                                + "%"
                            )
                    except (ValueError, TypeError):
                        pass
        self.wb.save(self.file_path)
        logger.success("excel was saved after inserting previous year block")
        subject = f"Отчет с данными за {date.strftime('%d.%m.%Y')}"
        body = f"Добавлены данные за прошлый год для даты {date.strftime('%d.%m.%Y')}.\nФайл во вложении."
        send_email_with_attachment(subject, body, self.file_path)

    def check_all_cities_filled(self, ws, data_row):
        for col in range(2, 7):
            all_metrics_filled = True
            for i in range(1, len(HEADERS) + 1):
                if ws.cell(row=data_row + i, column=col).value is None:
                    all_metrics_filled = False
                    break

            if not all_metrics_filled:
                return False

        return True

    def process_message(self, data: dict):
        try:
            date_str = data["Date"]
            city_key = data["City"]
            date = datetime.strptime(date_str, "%Y-%m-%d")
            city_name = CITY_MAPPING.get(city_key)

            if not city_name:
                logger.warning(f"⚠️ unknown city: {city_key}")
                return

            values = [
                int(data["ВСЕГО:"]),
                int(data["Потеряно:"]),
                int(data["Переведено:"]),
                int(data["Успешно завершено:"]),
                math.floor(
                    float(data["Клиенты, не дождавшиеся ответа, ждали в среднем:"])
                    + 0.5
                ),
                math.floor(float(data["В среднем клиенты ждут:"]) + 0.5),
                math.floor(float(data["В среднем разговор длится:"]) + 0.5),
                round((int(data["Потеряно:"]) / int(data["ВСЕГО:"])) * 100, 1)
                if int(data["ВСЕГО:"]) > 0
                else 0,
            ]

            sheet_name = f"{MONTH_NAMES[date.month]} {date.year}"

            if sheet_name not in self.wb.sheetnames:
                ws = self.create_new_sheet(sheet_name)
            else:
                ws = self.wb[sheet_name]

            logger.info(
                f"🔍 process data for {date_str} ({city_name}) on sheet {sheet_name}"
            )
            data_row = self.find_data_section(ws, date)

            if data_row is None:
                logger.info(f"⚠️ date {date_str} not found, create new section")
                data_row = self.create_new_data_section(ws, date)
                logger.success(f"✅ new section created at row: {data_row}")

            city_col = self.get_city_column(ws, data_row, city_name)

            if not city_col:
                logger.warning(f"⚠️ col for city '{city_name}' not found, adding new")
                # ищем последний занятый столбец в строке с датой
                last_col = 1
                for col in range(2, ws.max_column + 1):
                    if ws.cell(data_row, col).value:
                        last_col = col

                city_col = last_col + 1

                cell = ws.cell(data_row, city_col)
                cell.value = city_name
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = HEADER_ALIGNMENT
                cell.border = THICK_BORDER
                logger.info(
                    f"add new column {get_column_letter(city_col)} for city: {city_name}"
                )

            start_row = data_row + 1
            for i, value in enumerate(values):
                cell = ws.cell(row=start_row + i, column=city_col)
                self.format_data_cell(cell, value, i)

            logger.success(
                f"✅ data has been added {get_column_letter(city_col)}{start_row + i}"
            )
            self.wb.save(self.file_path)
            logger.success("💾 excel was saved")

            if self.check_all_cities_filled(ws, data_row):
                logger.info("all cities (B-F) filled for current date")

                prev_date = self.get_previous_year_same_weekday(date)
                prev_date_str = prev_date.strftime("%d.%m.%Y")
                weekday = WEEKDAYS[prev_date.weekday()]
                expected_header = f"{prev_date_str} ({weekday})"

                current_section_end = data_row + len(HEADERS)
                next_block_row = None

                for row in range(current_section_end + 1, ws.max_row + 1):
                    if ws.cell(row=row, column=1).value is not None:
                        next_block_row = row
                        break

                block_exists = False
                if next_block_row:
                    next_block_value = ws.cell(row=next_block_row, column=1).value
                    if next_block_value == expected_header:
                        block_exists = True
                        logger.info(
                            f"previous year block found at row {next_block_row}: {next_block_value}"
                        )

                if not block_exists:
                    logger.info("inserting previous year block")
                    self.insert_previous_year_block(ws, data_row, date)
                else:
                    logger.info(
                        "previous year block already exists, sending email anyway"
                    )
                    subject = f"Отчет с данными за {date.strftime('%d.%m.%Y')}"
                    body = f"Все данные заполнены для даты {date.strftime('%d.%m.%Y')}.\nФайл во вложении."
                    send_email_with_attachment(subject, body, self.file_path)
            else:
                logger.info(
                    "not all cities (B-F) filled yet, skipping previous year block"
                )

        except Exception as e:
            logger.error(f"❌ error: {str(e)}")
            import traceback

            traceback.print_exc()
